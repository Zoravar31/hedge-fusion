"""
HedgeFusion Data Exporter
============================
Export any HedgeFusion dataset to CSV, Excel (.xlsx), or JSON for
external use — tax filing, personal spreadsheets, sharing with a CA,
or importing into other tools.

Supported exports:
  trades      → paper_trades.csv reformatted (with computed P&L)
  holdings    → current config.py holdings with live prices
  portfolio   → latest portfolio_runner.py results
  watchlist   → current watchlist with latest scan data
  journal     → closed trade P&L detail from trade_journal
  memory      → agent_memory.json flattened to a table

Formats:
  csv   — always available (stdlib csv module)
  xlsx  — requires openpyxl (pip install openpyxl); falls back to
          csv with a warning if not installed
  json  — raw structured export

Usage:
    python data_exporter.py --what trades --format xlsx
    python data_exporter.py --what holdings --format csv
    python data_exporter.py --what journal --format xlsx --since 180
    python data_exporter.py --what all --format csv          # export everything
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from loguru import logger

load_dotenv(Path(__file__).parent / ".env")

OUTPUT_DIR = Path(__file__).parent / "outputs" / "exports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ──────────────────────────────────────────────
# Format writers
# ──────────────────────────────────────────────

def _write_csv(rows: list, path: Path) -> Path:
    if not rows:
        path.write_text("", encoding="utf-8")
        return path
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow(row)
    return path


def _write_xlsx(rows: list, path: Path, sheet_name: str = "Sheet1") -> Path:
    """Write rows to .xlsx using openpyxl. Falls back to CSV if not installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed — run: pip install openpyxl. Falling back to CSV.")
        csv_path = path.with_suffix(".csv")
        return _write_csv(rows, csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        wb.save(path)
        return path

    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Auto-size columns (approximate)
    for i, header in enumerate(headers, 1):
        max_len = max(
            [len(str(header))] + [len(str(row.get(header, ""))) for row in rows]
        )
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = min(max_len + 2, 40)

    wb.save(path)
    return path


def _write_json(data, path: Path) -> Path:
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    return path


def _export(rows: list, name: str, fmt: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fmt = fmt.lower()
    if fmt == "csv":
        path = OUTPUT_DIR / f"{name}_{ts}.csv"
        return _write_csv(rows, path)
    elif fmt == "xlsx":
        path = OUTPUT_DIR / f"{name}_{ts}.xlsx"
        return _write_xlsx(rows, path, sheet_name=name[:31])
    elif fmt == "json":
        path = OUTPUT_DIR / f"{name}_{ts}.json"
        return _write_json(rows, path)
    else:
        raise ValueError(f"Unknown format: {fmt}. Use csv, xlsx, or json.")


# ──────────────────────────────────────────────
# Dataset builders
# ──────────────────────────────────────────────

def export_trades(fmt: str = "csv", since_days: int = 365) -> Path:
    """Export raw paper trade log."""
    from trade_journal import load_paper_trades
    trades = load_paper_trades(since_days=since_days)
    rows = [
        {
            "date":             t["ts"].strftime("%Y-%m-%d %H:%M"),
            "order_id":         t.get("order_id", ""),
            "symbol":           t.get("symbol", ""),
            "type":             t.get("transaction_type", ""),
            "quantity":         t.get("quantity", 0),
            "fill_price":       t.get("fill_price", 0),
            "value_inr":        t.get("value_inr", 0),
            "stop_loss":        t.get("stop_loss", ""),
            "take_profit":      t.get("take_profit", ""),
            "status":           t.get("status", ""),
        }
        for t in trades
    ]
    return _export(rows, "trades", fmt)


def export_journal(fmt: str = "csv", since_days: int = 365) -> Path:
    """Export closed-trade P&L detail (matched buy/sell pairs)."""
    from trade_journal import load_paper_trades, compute_stats
    trades = load_paper_trades(since_days=since_days)
    stats  = compute_stats(trades)
    rows = [
        {
            "symbol":      c["symbol"],
            "entry_date":  c["entry_date"],
            "exit_date":   c["exit_date"],
            "entry_price": c["entry_price"],
            "exit_price":  c["exit_price"],
            "quantity":    c["quantity"],
            "pnl_inr":     c["pnl_inr"],
            "pnl_pct":     c["pnl_pct"],
            "won":         "YES" if c["won"] else "NO",
        }
        for c in stats.get("closed_trade_detail", [])
    ]
    return _export(rows, "journal", fmt)


def export_holdings(fmt: str = "csv") -> Path:
    """Export current config.py holdings with live prices and P&L."""
    from config import HOLDINGS
    from tools.india_data import get_nse_quote

    rows = []
    for h in HOLDINGS:
        ticker = h["ticker"]
        current_price = 0.0
        try:
            q = json.loads(get_nse_quote(ticker))
            current_price = (
                q.get("info", {}).get("currentPrice")
                or q.get("latest_close")
                or 0.0
            )
        except Exception:
            pass

        invested = h.get("qty", 0) * h.get("avg_buy_price", 0)
        current_value = h.get("qty", 0) * current_price
        pnl = current_value - invested if invested else 0

        rows.append({
            "ticker":         ticker,
            "sector":         h.get("sector", ""),
            "qty":            h.get("qty", 0),
            "avg_buy_price":  h.get("avg_buy_price", 0),
            "current_price":  round(float(current_price), 2),
            "invested_inr":   round(invested, 2),
            "current_value_inr": round(current_value, 2),
            "pnl_inr":        round(pnl, 2),
            "pnl_pct":        round(pnl / invested * 100, 2) if invested else 0,
            "stop_loss_pct":  h.get("stop_loss_pct", ""),
            "target_pct":     h.get("target_pct", ""),
        })
    return _export(rows, "holdings", fmt)


def export_watchlist(fmt: str = "csv") -> Path:
    """Export current watchlist."""
    from watchlist import load_watchlist
    wl = load_watchlist()
    rows = [
        {
            "ticker":       w["ticker"],
            "entry_target": w.get("entry_target", 0),
            "reason":       w.get("reason", ""),
            "added":        w.get("added", ""),
        }
        for w in wl
    ]
    return _export(rows, "watchlist", fmt)


def export_latest_portfolio(fmt: str = "csv") -> Path:
    """Export the most recent portfolio_runner.py JSON output as a flat table."""
    outputs_dir = Path(__file__).parent / "outputs"
    candidates = sorted(outputs_dir.glob("portfolio_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("No portfolio_*.json found. Run: python hf.py portfolio")

    data = json.loads(candidates[0].read_text(encoding="utf-8"))
    rows = []
    for r in data:
        rv = r.get("research_verdict", {})
        pm = r.get("pm_decision", {})
        ex = r.get("execution_result") or {}
        rows.append({
            "ticker":         r.get("ticker", ""),
            "sector":         r.get("holding_sector", ""),
            "recommendation": rv.get("recommendation", ""),
            "confidence":     rv.get("confidence", ""),
            "pm_decision":    pm.get("decision", ""),
            "entry_zone":     rv.get("entry_zone", ""),
            "stop_loss":      rv.get("stop_loss", ""),
            "target1":        rv.get("target1", ""),
            "risk_reward":    rv.get("risk_reward", ""),
            "order_status":   ex.get("order_id") or ex.get("status", ""),
        })
    return _export(rows, "portfolio", fmt)


def export_memory(fmt: str = "csv") -> Path:
    """Export agent_memory.json flattened to one row per verdict."""
    from agent_memory import _load_all
    all_mem = _load_all()
    rows = []
    for ticker, mem in all_mem.items():
        for h in mem.get("history", []):
            rows.append({
                "ticker":         ticker,
                "date":           h.get("date", ""),
                "recommendation": h.get("recommendation", ""),
                "pm_decision":    h.get("pm_decision", ""),
                "confidence":     h.get("confidence", ""),
                "entry_zone":     h.get("entry_zone", ""),
                "stop_loss":      h.get("stop_loss", ""),
                "target1":        h.get("target1", ""),
                "order_status":   h.get("order_status", ""),
            })
    return _export(rows, "memory", fmt)


# ──────────────────────────────────────────────
# Main dispatcher
# ──────────────────────────────────────────────

EXPORTERS = {
    "trades":    export_trades,
    "journal":   export_journal,
    "holdings":  export_holdings,
    "watchlist": export_watchlist,
    "portfolio": export_latest_portfolio,
    "memory":    export_memory,
}


def run_export(what: str, fmt: str = "csv", since_days: int = 365) -> list:
    """
    Export one dataset, or "all" for every dataset.
    Returns list of Path objects written.
    """
    if what == "all":
        targets = list(EXPORTERS.keys())
    else:
        targets = [what]

    written = []
    for name in targets:
        fn = EXPORTERS.get(name)
        if not fn:
            print(f"  ⚠ Unknown export target: {name}")
            continue
        try:
            if name in ("trades", "journal"):
                path = fn(fmt, since_days)
            else:
                path = fn(fmt)
            written.append(path)
            size_kb = path.stat().st_size / 1024 if path.exists() else 0
            print(f"  ✅ {name:<12} → {path.name}  ({size_kb:.1f} KB)")
        except Exception as e:
            print(f"  ❌ {name:<12} failed: {e}")
            logger.error("export {} failed: {}", name, e)

    return written


def main():
    import argparse
    parser = argparse.ArgumentParser(description="HedgeFusion Data Exporter")
    parser.add_argument("--what",   default="all",
                        choices=list(EXPORTERS.keys()) + ["all"],
                        help="Dataset to export (default: all)")
    parser.add_argument("--format", default="csv", choices=["csv", "xlsx", "json"])
    parser.add_argument("--since",  type=int, default=365, help="Days of history for trades/journal")
    args = parser.parse_args()

    print(f"\n{'━'*55}")
    print(f"  HEDGEFUSION DATA EXPORTER")
    print(f"  Target: {args.what} | Format: {args.format}")
    print(f"{'━'*55}\n")

    written = run_export(args.what, args.format, args.since)

    print(f"\n{'━'*55}")
    print(f"  ✅ {len(written)} file(s) exported to outputs/exports/")
    print(f"{'━'*55}\n")


if __name__ == "__main__":
    main()
